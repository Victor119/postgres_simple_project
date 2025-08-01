import base64
import calendar
from datetime import datetime, date, timedelta
from flask import Flask, jsonify, render_template, render_template_string, request, redirect, session, url_for 
from flask_sqlalchemy import SQLAlchemy
import msal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from pypdf import PdfReader, PdfWriter
import pandas as pd
import requests
import shutil
from typing import List, Dict
import zipfile

app = Flask(__name__)

# Foloseste un secret key aleator la fiecare pornire, astfel incat sesiunile vechi sa fie invalidate
app.secret_key = os.environ.get('SECRET_KEY') or os.urandom(24)

#DB configuration
app.config['SQLALCHEMY_DATABASE_URI'] = (os.environ.get('DATABASE_URL') or 'postgresql://postgres:postgres@localhost:5432/master')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ----------------------------- HTML TEMPLATE ----------------------------------

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{{ title }}</title>
    <style>
        body {
            margin: 0;
            padding: 0;
            background-color: #f0f0f0;
        }
        .window {
            position: absolute;
            left: {{ x }}px;
            top: {{ y }}px;
            width: {{ w }}px;
            height: {{ h }}px;
            border: 2px solid #333;
            background-color: #ccc;
            box-shadow: 2px 2px 10px rgba(0,0,0,0.5);
            padding: 10px;
        }
        .display-container {
            position: absolute;
            top: 50px;
            left: 100px;
            display: flex;
            align-items: center;
        }
        .display-label {
            font-family: sans-serif;
            font-size: 14px;
            margin-right: 10px;
        }
        .display-box {
            width: 200px;
            height: 50px;
            border: 1px solid #666;
            background-color: #fff;
            font-family: monospace;
            font-size: 14px;
            padding: 5px;
            box-sizing: border-box;
        }
        .radio-button {
            position: absolute;
            top: 120px;
            left: 100px;
            font-family: sans-serif;
            font-size: 14px;
        }
        .return-button {
            position: absolute;
            left: {{ ret_x }}px;
            top: {{ ret_y }}px;
            width: {{ ret_w }}px;
            height: {{ ret_h }}px;
            font-size: 12px;
        }
        .radio-group {
            position: absolute;
            top: 150px;
            left: 160px;
            width: 200px;
            border: 1px solid #333;
            background-color: #eee;
            padding: 10px;
        }
    </style>
    <script>
        // Enforce max 256 characters per line and allow Enter key
        document.addEventListener('DOMContentLoaded', function() {
            function enforceMaxLineLength(textarea) {
                const lines = textarea.value.split('\n');
                for (let i = 0; i < lines.length; i++) {
                    if (lines[i].length > 256) {
                        lines[i] = lines[i].substring(0, 256);
                    }
                }
                textarea.value = lines.join('\n');
            }
            ['edit_box', 'edit_box2'].forEach(function(id) {
                const ta = document.getElementById(id);
                if (ta) {
                    ta.addEventListener('input', function() { enforceMaxLineLength(ta); });
                }
            });
        });
    </script>
</head>
<body>
    <div class="window">
        <div class="display-container">
            <div class="display-label">export Excel</div>
            <input class="display-box" type="text" value="{{ first_text }}" readonly>

            <div class="display-label" style="margin-left: 20px;">export PDF</div>
            <input class="display-box" type="text" value="{{ second_text }}" readonly>

            <div class="display-label" style="margin-left: 20px;">send Excel</div>
            <input class="display-box" type="text" value="{{ third_text }}" readonly>
            
            <div class="display-label" style="margin-left: 20px;">send PDF</div>
            <input class="display-box" type="text" value="{{ fourth_text }}" readonly>
        </div>
        
        <form method="post">
            <div class="radio-group">
                {% set custom_labels = ["export Excel", "export PDF", "send Excel", "send PDF"] %}
                {% for i in range(radio_buttons | length) %}
                <div class="radio-option">
                    <input type="radio" id="{{ radio_buttons[i].id }}" name="radio_option" value="{{ i + 1 }}" 
                        {% if selected_choice == i + 1 %}checked{% endif %}>
                    <label for="{{ radio_buttons[i].id }}">{{ custom_labels[i] }}</label>
                </div>
                {% endfor %}
            </div>
            
            <div style="position: absolute; top: 130px; left: 400px;">
                <label for="edit_box" style="font-family: sans-serif; font-size: 14px;">My Input</label><br>
                <textarea id="edit_box" name="edit_box" style="width: 150px; height: 100px; font-family: monospace; font-size: 14px;">{{ current_input }}</textarea>
            </div>
            
            <div style="position: absolute; top: 130px; left: 620px;">
                <label for="edit_box2" style="font-family: sans-serif; font-size: 14px;">My Input&nbsp;2</label><br>
                <textarea id="edit_box2" name="edit_box2" style="width: 150px; height: 100px; font-family: monospace; font-size: 14px;">{{ second_input }}</textarea>
            </div>
            
            <button class="return-button" type="submit">&Return</button>
        </form>
    </div>
</body>
</html>
"""

# ----------------------------- DATABASE Models ---------------------

# Employee model conform structurii existente
class Employee(db.Model):
    __tablename__ = 'employees'

    employee_id = db.Column(db.Integer,
                            primary_key=True)
    manager_id  = db.Column(db.Integer,
                            db.ForeignKey('employees.employee_id'),
                            nullable=True)
    role        = db.Column(db.String,      # 'manager' sau 'employee'
                            nullable=False)
    first_name  = db.Column(db.String(100))
    last_name   = db.Column(db.String(100))
    cnp         = db.Column(db.String(20),
                            unique=True,
                            nullable=False)
    username    = db.Column(db.String(64),
                            unique=True,
                            nullable=False)
    password    = db.Column(db.String(128),
                            nullable=False)
    email       = db.Column(db.String(120),
                            unique=True,
                            nullable=False)
    address     = db.Column(db.String(200))
    city        = db.Column(db.String(100))
    country     = db.Column(db.String(100))
    created_at  = db.Column(db.DateTime)

    @property
    def full_name(self):
        """Returnează numele complet al angajatului"""
        if self.first_name and self.last_name:
            return f"{self.first_name} {self.last_name}"
        elif self.first_name:
            return self.first_name
        elif self.last_name:
            return self.last_name
        else:
            return self.username or "N/A"

class Salary(db.Model):
    __tablename__ = 'salaries'
    salary_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    base_salary = db.Column(db.Numeric(10, 2))
    month = db.Column(db.Date)

class Bonus(db.Model):
    __tablename__ = 'bonuses'
    bonus_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    amount = db.Column(db.Numeric(10, 2))
    description = db.Column(db.String(255))
    month = db.Column(db.Date)

class Vacation(db.Model):
    __tablename__ = 'vacations'
    vacation_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    number_of_days = db.Column(db.Integer)
    reason = db.Column(db.String(255))

class WorkDay(db.Model):
    __tablename__ = 'work_days'
    work_day_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    month = db.Column(db.Date)
    number_of_days = db.Column(db.Integer)

class ArchivedFile(db.Model):
    __tablename__ = 'archived_files'
    file_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    file_name = db.Column(db.String(255))
    file_type = db.Column(db.String(50))
    path = db.Column(db.String(500))
    sent_date = db.Column(db.Date)

# ----------------------------- HELPER FUNCTIONS FOR EXCEL ---------------------

def get_current_month_start():
    """Returneaza primul zi din luna curenta"""
    today = date.today()
    return date(today.year, today.month, 1)

def get_current_month_end():
    """ReturneazA ultima zi din luna curentA"""
    today = date.today()
    last_day = calendar.monthrange(today.year, today.month)[1]
    return date(today.year, today.month, last_day)

def get_previous_month_range():
    """
    Returnează primul și ultimul day din luna precedentă față de data curentă.
    Exemplu: azi este 2025-08-01, va returna (2025-07-01, 2025-07-31)
    """
    today = date.today()
    # 1) prima zi din luna curentă
    first_of_this = date(today.year, today.month, 1)
    # 2) ultima zi din luna precedentă = prima zi din luna curentă minus o zi
    last_of_prev = first_of_this - timedelta(days=1)
    # 3) prima zi din luna precedentă
    first_of_prev = date(last_of_prev.year, last_of_prev.month, 1)
    return first_of_prev, last_of_prev

def get_employee_data_for_excel(email_addresses):
    """
    Colectează datele pentru toți angajații cu adresele specificate
    pentru luna precedentă față de data curentă.
    """
    try:
        employees_data = []
        start, end = get_previous_month_range()
        
        # Dacă nu s-au dat email-uri, includem toți angajații
        if not email_addresses:
            email_addresses = [
                emp.email
                for emp in Employee.query.with_entities(Employee.email).all()
            ]
        
        for email in email_addresses:
            email = email.strip()
            if not email:
                continue
                
            # găsim angajatul
            employee = Employee.query.filter_by(email=email).first()
            if not employee:
                continue
            
            # salariul pentru luna precedentă
            salary = Salary.query.filter(
                Salary.employee_id == employee.employee_id,
                Salary.month >= start,
                Salary.month <= end
            ).first()
            
            # zile lucrate pentru luna precedentă
            work_days = WorkDay.query.filter(
                WorkDay.employee_id == employee.employee_id,
                WorkDay.month >= start,
                WorkDay.month <= end
            ).first()
            
            # zile de concediu în luna precedentă
            vacation_days = db.session.query(db.func.sum(Vacation.number_of_days)).filter(
                Vacation.employee_id == employee.employee_id,
                Vacation.start_date >= start,
                Vacation.end_date <= end
            ).scalar() or 0
            
            # bonusuri în luna precedentă
            bonuses = Bonus.query.filter(
                Bonus.employee_id == employee.employee_id,
                Bonus.month >= start,
                Bonus.month <= end
            ).all()
            
            # pregătim structura pentru Excel
            employee_info = {
                'name': employee.full_name,
                'email': employee.email,
                'salary': float(salary.base_salary) if salary and salary.base_salary else 0.0,
                'work_days': work_days.number_of_days if work_days else 0,
                'vacation_days': vacation_days,
                'bonuses': []
            }
            
            for bonus in bonuses:
                employee_info['bonuses'].append({
                    'amount': float(bonus.amount) if bonus.amount else 0.0,
                    'description': bonus.description or 'N/A'
                })
            
            employees_data.append(employee_info)
        
        return employees_data
    
    except Exception as e:
        print(f"Error collecting employee data for Excel: {e}")
        return []

def create_excel_file(employees_data, output_path):
    """
    Creează fișierul Excel cu datele angajaților
    """
    try:
        # Create the workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Summary"
        
        # Styles for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # The headers of the columns
        headers = [
            "Employee Name", 
            "Email", 
            "Base Salary", 
            "Working Days", 
            "Vacation Days", 
            "Total Bonuses", 
            "Bonus Details"
        ]
        
        # Add the headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add the employee data
        for row, employee in enumerate(employees_data, 2):
            ws.cell(row=row, column=1, value=employee['name'])
            ws.cell(row=row, column=2, value=employee['email'])
            ws.cell(row=row, column=3, value=employee['salary'])
            ws.cell(row=row, column=4, value=employee['work_days'])
            ws.cell(row=row, column=5, value=employee['vacation_days'])
            
            # Calculate the total bonuses
            total_bonuses = sum(bonus['amount'] for bonus in employee['bonuses'])
            ws.cell(row=row, column=6, value=total_bonuses)
            
            # Add the details of the bonuses
            bonus_details = "; ".join([f"{bonus['description']}: {bonus['amount']}" for bonus in employee['bonuses']])
            ws.cell(row=row, column=7, value=bonus_details if bonus_details else "No bonuses")
        
        # Adjust the width of the columns
        column_widths = [20, 25, 15, 15, 15, 15, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Save the file
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False

# ----------------------------- HELPER FUNCTIONS FOR PDF ---------------------

def get_employee_data_for_pdf(email_addresses):
    """
    Colectează datele necesare pentru exportul PDF
    (doar detalii personale și salariul datorat)
    pentru luna precedentă față de data curentă.
    """
    try:
        employees_data = []
        start, end = get_previous_month_range()

        for email in (e.strip() for e in email_addresses):
            if not email:
                continue

            employee = Employee.query.filter_by(email=email).first()
            if not employee:
                continue

            salary_rec = Salary.query.filter(
                Salary.employee_id == employee.employee_id,
                Salary.month >= start,
                Salary.month <= end
            ).first()

            employees_data.append({
                "employee_id": employee.employee_id,
                "first_name":  employee.first_name or "",
                "last_name":   employee.last_name or "",
                "cnp":         employee.cnp or "",
                "salary":      float(salary_rec.base_salary) if salary_rec and salary_rec.base_salary else 0.0,
            })

        return employees_data
    except Exception as e:
        print(f"Error collecting PDF data: {e}")
        return []

def create_pdf_file(employees_data, output_path):
    """
    Creates one aggregated PDF containing one short section
    for every employee in `employees_data`.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen       import canvas

        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        y = height - 40

        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, y, "Monthly Salary Report")
        y -= 30

        for emp in employees_data:
            if y < 120:          # simple page‑break
                c.showPage()
                y = height - 40
                c.setFont("Helvetica-Bold", 16)
                c.drawString(40, y, "Monthly Salary Report (cont.)")
                y -= 30

            c.setFont("Helvetica-Bold", 12)
            full_name = f"{emp['first_name']} {emp['last_name']}".strip()
            c.drawString(40, y, f"Employee: {full_name}")
            y -= 15

            c.setFont("Helvetica", 11)
            c.drawString(60, y, f"Employee ID: {emp['employee_id']}")
            y -= 15
            c.drawString(60, y, f"CNP: {emp['cnp']}")
            y -= 15
            c.drawString(60, y, f"Salary for current month: {emp['salary']:.2f}")
            y -= 25

        c.save()
        return True
    except Exception as e:
        print(f"Error creating PDF file: {e}")
        return False

# ----------------------------- EMAIL HELPER FUNCTION ---------------------

CONFIG_PATH = "C:\\Users\\virosca\\Documents\\programs_python\\tema2\\app\\azure_credentials.txt"

def load_config(path: str = CONFIG_PATH) -> dict:
    """
    Incarca chei=valori dintr-un fisier text.
    Liniile goale si comentariile (incepând cu '#') sunt ignorate.
    """
    config = {}
    if not os.path.exists(path):
        raise FileNotFoundError(f"Config file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            key, val = line.split("=", 1)
            config[key.strip()] = val.strip()
    return config

# Incarc o singura data la import:
_cfg = load_config()
client_id = _cfg.get("client_id")
tenant_id = _cfg.get("tenant_id")

def send_file_via_email(
    file_path: str,
    recipient_emails: List[str] 
) -> bool:
    """
    Trimite email folosind delegated permissions - nu necesită admin consent
    Utilizatorul se va autentifica interactiv în browser
    """
    
    #1. Configure the MSAL application for delegated permissions
    app = msal.PublicClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}"
    )
    
    # Goals for delegated permissions (do not require admin consent)
    scopes = ["https://graph.microsoft.com/Mail.Send"]
    
    # 2. Try to get the token from the cache.
    accounts = app.get_accounts()
    result = None
    
    if accounts:
        #Try to obtain a silent token for the first account.
        result = app.acquire_token_silent(scopes, account=accounts[0])
    
    if not result:
        #If there is no token in the cache, request interactive authentication.
        print("Se deschide browser-ul pentru autentificare...")
        result = app.acquire_token_interactive(scopes)
    
    if "access_token" not in result:
        print(f"Eroare la autentificare: {result.get('error_description')}")
        return False
    
    token = result["access_token"]
    print("Autentificare reusita!")
    
    # 3. Read the file
    try:
        with open(file_path, 'rb') as f:
            file_content = f.read()
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        filename = os.path.basename(file_path)
    except Exception as ex:
        print(f"Eroare la citirea fișierului: {ex}")
        return False
    
    # 4. Build the payload
    recipients = [{"emailAddress": {"address": email}} for email in recipient_emails]
    
    email_payload = {
        "message": {
            "subject": "Employee Summary Report",
            "body": {
                "contentType": "HTML",
                "content": """
                <p>Hello,</p>
                <p>Please find attached the <strong>Employee Summary Report</strong>.</p>
                <p>Best regards,<br/>HR Department</p>
                """
            },
            "toRecipients": recipients,
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": filename,
                    "contentType": "application/octet-stream",
                    "contentBytes": file_base64
                }
            ]
        }
    }
    
    # 5. Send the email
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    
    graph_url = "https://graph.microsoft.com/v1.0/me/sendMail"
    
    try:
        response = requests.post(graph_url, json=email_payload, headers=headers)
        
        if response.status_code == 202:
            print("Email trimis cu succes!")
            return True
        else:
            print(f"Eroare: {response.status_code}")
            print(f"Răspuns: {response.text}")
            return False
            
    except Exception as e:
        print(f"Eroare la trimiterea email-ului: {e}")
        return False

# ----------------------------- Archives helper functions --------------------------------
_pending_excel_for_managers = {}
_manager_archives_in_progress = {}  # tracker pentru arhivele in progres
_manager_expected_employees = {}  # manager_email -> set of employee_emails

def create_archive_for_manager(manager_email, excel_path, pdf_files):
    """
    Creeaza arhiva finala cu Excel + toate PDF-urile (manager + employees)
    si o muta in folderul archives_final
    """
    try:
        # Folderul temporar pentru arhive
        temp_archive_folder = "C:\\Users\\virosca\\Documents\\programs_python\\tema2\\app\\archives"
        # Folderul final pentru arhive complete
        final_archive_folder = "C:\\Users\\virosca\\Documents\\programs_python\\tema2\\app\\archives_final"
        
        os.makedirs(temp_archive_folder, exist_ok=True)
        os.makedirs(final_archive_folder, exist_ok=True)
        
        # Gaseste managerul
        manager = Employee.query.filter_by(email=manager_email).first()
        if not manager:
            print(f"Manager not found: {manager_email}")
            return False
        
        # Creeaza numele arhivei
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_name = f"manager_{manager.employee_id}_{timestamp}.zip"
        temp_archive_path = os.path.join(temp_archive_folder, archive_name)
        final_archive_path = os.path.join(final_archive_folder, archive_name)
        
        # Lista fisierelor de inclus in arhiva
        files_to_archive = []
        
        # Adauga Excel-ul
        if excel_path and os.path.exists(excel_path):
            files_to_archive.append(excel_path)
            print(f"Added Excel to archive: {excel_path}")
        
        # Adauga toate PDF-urile
        for pdf_file in pdf_files:
            if os.path.exists(pdf_file):
                files_to_archive.append(pdf_file)
                print(f"Added PDF to archive: {pdf_file}")
        
        if not files_to_archive:
            print("No files to archive")
            return False
        
        # Creeaza arhiva ZIP temporara
        with zipfile.ZipFile(temp_archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in files_to_archive:
                # Foloseste doar numele fisierului in arhiva (fara path-ul complet)
                arcname = os.path.basename(file_path)
                zipf.write(file_path, arcname)
                print(f"Archived: {file_path} as {arcname}")
        
        # Muta arhiva in folderul final
        shutil.move(temp_archive_path, final_archive_path)
        print(f"Archive moved to final location: {final_archive_path}")
        
        return True
        
    except Exception as e:
        print(f"Error creating archive: {e}")
        return False

def init_manager_archive_with_employees(manager_email, excel_path, expected_employee_emails):
    """
    Initializeaza o arhiva pentru manager si inregistreaza lista de angajati asteptati
    """
    try:
        global _manager_archives_in_progress, _manager_expected_employees
        
        # Folderul temporar pentru arhive
        temp_archive_folder = "C:\\Users\\virosca\\Documents\\programs_python\\tema2\\app\\archives"
        os.makedirs(temp_archive_folder, exist_ok=True)
        
        # Gaseste managerul
        manager = Employee.query.filter_by(email=manager_email).first()
        if not manager:
            print(f"Manager not found: {manager_email}")
            return False
        
        # Creează numele arhivei
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_name = f"manager_{manager.employee_id}_{timestamp}.zip"
        temp_archive_path = os.path.join(temp_archive_folder, archive_name)
        
        # Creeaza arhiva initiala cu Excel-ul
        with zipfile.ZipFile(temp_archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            if excel_path and os.path.exists(excel_path):
                arcname = os.path.basename(excel_path)
                zipf.write(excel_path, arcname)
                print(f"Initialized archive with Excel: {excel_path}")
        
        # Salveaza informatia despre arhiva in progres
        _manager_archives_in_progress[manager_email] = {
            'archive_path': temp_archive_path,
            'manager_id': manager.employee_id,
            'timestamp': datetime.now(),
            'pdf_count': 0,
            'received_pdfs': set()  # email-urile care au primit deja PDF-uri
        }
        
        # Inregistreaza angajatii așteptati pentru acest manager
        _manager_expected_employees[manager_email] = set(expected_employee_emails)
        
        print(f"Archive initialized for manager {manager_email}: {temp_archive_path}")
        print(f"Expected employees: {expected_employee_emails}")
        return True
        
    except Exception as e:
        print(f"Error initializing manager archive: {e}")
        return False


def add_pdf_to_manager_archive(manager_email, pdf_path):
    """
    Adauga un PDF la arhiva managerului (in progres)
    """
    try:
        global _manager_archives_in_progress
        
        if manager_email not in _manager_archives_in_progress:
            print(f"No archive in progress for manager: {manager_email}")
            return False
        
        archive_info = _manager_archives_in_progress[manager_email]
        archive_path = archive_info['archive_path']
        
        if not os.path.exists(archive_path):
            print(f"Archive file not found: {archive_path}")
            return False
        
        if not os.path.exists(pdf_path):
            print(f"PDF file not found: {pdf_path}")
            return False
        
        # Citeste arhiva existenta si adauga noul PDF
        temp_path = f"{archive_path}.tmp"
        with zipfile.ZipFile(archive_path, 'r') as source_zip:
            with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as target_zip:
                # Copiaza toate fisierele existente
                for item in source_zip.infolist():
                    data = source_zip.read(item.filename)
                    target_zip.writestr(item, data)
                
                # Adauga noul PDF
                arcname = os.path.basename(pdf_path)
                target_zip.write(pdf_path, arcname)
                print(f"Added PDF to archive: {pdf_path}")
        
        # Inlocuieste arhiva originala cu cea actualizata
        os.replace(temp_path, archive_path)
        
        # Actualizeaza contorul
        archive_info['pdf_count'] += 1
        
        return True
        
    except Exception as e:
        print(f"Error adding PDF to manager archive: {e}")
        return False

def finalize_manager_archive(manager_email):
    """
    Finalizeaza arhiva managerului si o muta in folderul archives_final
    """
    try:
        global _manager_archives_in_progress
        
        if manager_email not in _manager_archives_in_progress:
            print(f"No archive in progress for manager: {manager_email}")
            return False
        
        archive_info = _manager_archives_in_progress[manager_email]
        temp_archive_path = archive_info['archive_path']
        
        # Folderul final pentru arhive complete
        final_archive_folder = "C:\\Users\\virosca\\Documents\\programs_python\\tema2\\app\\archives_final"
        os.makedirs(final_archive_folder, exist_ok=True)
        
        # Calculeaza path-ul final
        archive_name = os.path.basename(temp_archive_path)
        final_archive_path = os.path.join(final_archive_folder, archive_name)
        
        # Muta arhiva din archives in archives_final
        shutil.move(temp_archive_path, final_archive_path)
        
        print(f"Archive finalized and moved to: {final_archive_path}")
        print(f"Total PDFs in archive: {archive_info['pdf_count']}")
        
        # Curata din tracking
        del _manager_archives_in_progress[manager_email]
        
        return True
        
    except Exception as e:
        print(f"Error finalizing manager archive: {e}")
        return False

def check_and_create_archive_for_manager(manager_email, excel_path):
    """
    Inregistreaza Excel-ul si determina toti angajatii care vor primi PDF-uri
    """
    try:
        global _pending_excel_for_managers
        
        # Gaseste managerul
        manager = Employee.query.filter_by(email=manager_email).first()
        if not manager:
            print(f"Manager not found: {manager_email}")
            return
        
        # Gaseste toti angajatii sub acest manager
        manager_employees = Employee.query.filter_by(manager_id=manager.employee_id).all()
        expected_emails = [emp.email for emp in manager_employees] + [manager_email]  # include managerul
        
        print(f"Manager {manager_email} expects PDFs for: {expected_emails}")
        
        _pending_excel_for_managers[manager_email] = {
            'excel_path': excel_path,
            'timestamp': datetime.now(),
            'expected_employees': set(expected_emails)
        }
        
        # Initializeaza arhiva cu lista completa de angajati asteptati
        init_manager_archive_with_employees(manager_email, excel_path, expected_emails)
        
    except Exception as e:
        print(f"Error in check_and_create_archive_for_manager_complete: {e}")

def check_and_create_archive_for_manager_with_pdfs(manager_email, pdf_files):
    """
    Verifica si construieste progresiv arhiva cand s-au trimis PDF-urile
    Finalizeaza arhiva doar cand toate PDF-urile au fost adaugate
    """
    try:
        global _pending_excel_for_managers, _manager_archives_in_progress
        
        # Verifica daca exista Excel trimis recent pentru acest manager
        if manager_email in _pending_excel_for_managers:
            excel_info = _pending_excel_for_managers[manager_email]
            
            # Verifica daca Excel-ul a fost trimis in ultimele 10 minute
            time_diff = datetime.now() - excel_info['timestamp']
            if time_diff.total_seconds() <= 600:
                
                # Gaseste managerul
                manager = Employee.query.filter_by(email=manager_email).first()
                if not manager:
                    print(f"Manager not found: {manager_email}")
                    return
                
                print(f"Processing archive for manager {manager_email}")
                print(f"PDF files to add: {len(pdf_files)}")
                
                # Verifica daca arhiva este in progres
                if manager_email in _manager_archives_in_progress:
                    # Adauga toate PDF-urile la arhiva in progres
                    successful_adds = 0
                    for pdf_file in pdf_files:
                        if os.path.exists(pdf_file):
                            if add_pdf_to_manager_archive(manager_email, pdf_file):
                                successful_adds += 1
                                print(f"  Successfully added: {os.path.basename(pdf_file)}")
                            else:
                                print(f"  Failed to add: {os.path.basename(pdf_file)}")
                        else:
                            print(f"  PDF not found: {pdf_file}")
                    
                    print(f"Successfully added {successful_adds}/{len(pdf_files)} PDFs to archive")
                    
                    # Finalizeaza arhiva (muta in archives_final)
                    if finalize_manager_archive(manager_email):
                        print(f"Archive successfully finalized for manager {manager_email}")
                        # Curata din pending
                        del _pending_excel_for_managers[manager_email]
                    else:
                        print(f"Failed to finalize archive for manager {manager_email}")
                        
                else:
                    print(f"No archive in progress found for manager {manager_email}")
                    
            else:
                print(f"Excel for manager {manager_email} is too old ({time_diff.total_seconds()} seconds)")
                # Curata intrarea expirata
                del _pending_excel_for_managers[manager_email]
        else:
            print(f"No pending Excel found for manager {manager_email}")
            
    except Exception as e:
        print(f"Error in check_and_create_archive_for_manager_with_pdfs: {e}")

# ----------------------------- API ENDPOINTS HELPER FUNCTIONS --------------------------------

def export_excel(output_folder: str, email_addresses: List[str]) -> str:
    """Returneaza calea completa a fisierului .xlsx creat sau ridica exceptie."""
    os.makedirs(output_folder, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"employee_summary_{timestamp}.xlsx"
    output_path = os.path.join(output_folder, filename)

    employees_data = get_employee_data_for_excel(email_addresses)
    if not employees_data:
        raise ValueError("No employee data found")
    if not create_excel_file(employees_data, output_path):
        raise RuntimeError("Failed to create Excel file")

    return output_path


def export_pdf(output_folder: str, email_addresses: List[str]) -> str:
    """Returneaza calea completa a fisierului .pdf creat sau ridica exceptie."""
    os.makedirs(output_folder, exist_ok=True)
    data = get_employee_data_for_pdf(email_addresses)
    if not data:
        raise ValueError("No employee data found")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"employee_summary_{timestamp}.pdf"
    output_path = os.path.join(output_folder, filename)

    if not create_pdf_file(data, output_path):
        raise RuntimeError("Failed to create PDF file")

    return output_path


def send_excel(file_path: str, recipient_emails: List[str]) -> None:
    """Trimite Excel-ul si, in caz ca recipientul e manager, initializeaza arhiva."""
    if not send_file_via_email(file_path, recipient_emails):
        raise RuntimeError("Failed to send Excel")
    # salvare in arhiva
    for email in recipient_emails:
        emp = Employee.query.filter_by(email=email).first()
        db.session.add(ArchivedFile(
            employee_id=emp.employee_id if emp else None,
            file_name=os.path.basename(file_path),
            file_type='Excel',
            path=file_path,
            sent_date=date.today()
        ))
        # daca e manager, initializeaza arhiva
        if emp and emp.role == 'manager':
            check_and_create_archive_for_manager(email, file_path)
    db.session.commit()


def send_pdfs(file_path: str, recipient_emails: List[str]) -> Dict[str, str]:
    """
    Cripteaza cu CNP pentru fiecare angajat,
    le trimite si le arhivează; returneaza un dict email->status.
    """
    results = {}
    for email in recipient_emails:
        emp = Employee.query.filter_by(email=email).first()
        if not emp or not emp.cnp:
            results[email] = 'Employee or CNP not found'
            continue

        # creeaza fisier criptat
        reader = PdfReader(file_path)
        writer = PdfWriter()
        for p in reader.pages:
            writer.add_page(p)
        # Folosim parametrul corect user_password (si, optional, owner_password)
        writer.encrypt(
            user_password=emp.cnp,
            owner_password=emp.cnp,
            use_128bit=True
        )

        # genereaza numele si calea fisierului criptat
        base = os.path.splitext(os.path.basename(file_path))[0]
        enc_name = f"{base}_{emp.employee_id}.pdf"
        enc_path = os.path.join(os.path.dirname(file_path), enc_name)
        with open(enc_path, 'wb') as out_f:
            writer.write(out_f)

        # trimite prin Graph API
        sent = send_file_via_email(enc_path, [email])
        results[email] = 'sent' if sent else 'failed'

        if sent:
            # salveaza in arhiva
            db.session.add(ArchivedFile(
                employee_id=emp.employee_id,
                file_name=enc_name,
                file_type='PDF',
                path=enc_path,
                sent_date=date.today()
            ))
            # adauga PDF-ul in arhivele managerilor relevanti
            Controller().add_pdf_to_relevant_manager_archives(email, enc_path)

    db.session.commit()
    # incearca sa finalizeze arhivele complete
    Controller().check_and_finalize_complete_archives()
    return results

# ----------------------------- API ENDPOINTS --------------------------------

@app.route("/createAggregatedEmployeeData", methods=["POST"])
def create_aggregated_employee_data():
    data = request.get_json() or {}
    folder = data.get('output_folder', 'app/generated_excel')
    emails = data.get('emails') or []
    try:
        path = export_excel(folder, emails)
        return jsonify({
            'success': True,
            'filename': os.path.basename(path),
            'path': path,
            'message': f'{os.path.basename(path)} created successfully'
        }), 200
    except ValueError as ve:
        return jsonify({'error': str(ve)}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/createPdfForEmployees", methods=["POST"])
def create_pdf_for_employees():
    data = request.get_json() or {}
    folder = data.get('output_folder', 'app/generated_pdf')
    emails = data.get('emails') or []
    try:
        path = export_pdf(folder, emails)
        return jsonify({
            'success': True,
            'filename': os.path.basename(path),
            'path': path,
            'message': f'{os.path.basename(path)} created successfully'
        }), 200
    except ValueError as ve:
        return jsonify({'error': str(ve)}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/sendAggregatedEmployeeData", methods=["POST"])
def send_aggregated_employee_data():
    data = request.get_json() or {}
    file_path = data.get('file_path', '').strip()
    emails = data.get('emails') or []
    if not file_path:
        return jsonify({'error': 'No file path provided'}), 400
    if not os.path.exists(file_path):
        return jsonify({'error': f'File not found: {file_path}'}), 404

    try:
        send_excel(file_path, emails)
        return jsonify({
            'success': True,
            'message': f'Excel file sent successfully to {len(emails)} recipients',
            'file_path': file_path,
            'recipients': emails
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/sendPdfToEmployees", methods=["POST"])
def send_pdf_to_employees():
    data = request.get_json() or {}
    file_path = data.get('file_path', '').strip()
    emails = data.get('emails') or []
    if not file_path:
        return jsonify({'error': 'No file path provided'}), 400
    if not file_path.lower().endswith('.pdf'):
        return jsonify({'error': 'Provided file is not a PDF'}), 400
    if not os.path.exists(file_path):
        return jsonify({'error': f'File not found: {file_path}'}), 404

    try:
        results = send_pdfs(file_path, emails)
        return jsonify({'results': results}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        # Cautam direct in tabela employees
        user = Employee.query.filter_by(
            username=username,
            password=password
        ).first()

        if user:
            session["user"] = user.username
            return redirect(url_for("index"))
        else:
            error = "Credentiale invalide"
            return render_template("login.html", error=error)

    # GET: afisam formularul
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("user", None)     # sterge cheia din sesiune
    return redirect(url_for("login"))

# ----------------------------- CLASSES ---------------------------------------

class Point:
    def __init__(self, x, y):
        self._x = x
        self._y = y

    def getX(self):
        return self._x

    def getY(self):
        return self._y
    
    def setY(self, y):
        self._y = y

class Fl_Output:
    def __init__(self, x, y, w, h, label=None):
        self.x = x
        self.y = y
        self.width = w
        self.height = h
        self.label = label
        self._value = ""

    def value(self, txt):
        self._value = txt

    def redraw(self):
        pass

    def getText(self):
        return self._value

class MyDisplayBox(Fl_Output):
    def __init__(self, pos: Point, w: int, h: int, label: str = None):
        super().__init__(pos.getX(), pos.getY(), w, h, label)

    def setText(self, txt: str):
        self.value(txt)
        self.redraw()

class MyReturnButton:
    def __init__(self, pos: Point, w: int, h: int, label: str = "&Return"):
        self.x = pos.getX()
        self.y = pos.getY()
        self.w = w
        self.h = h
        self.label = label
        self.tooltip = "Push Return button to exit"
        self.labelsize = 12

    def getRenderParams(self):
        return {
            "ret_x": self.x,
            "ret_y": self.y,
            "ret_w": self.w,
            "ret_h": self.h,
            "label": self.label
        }

# ----------------------------- EDIT BOX CLASS ----------------------------------

class MyEditBox:
    def __init__(self, pos: Point, w: int, h: int, label: str):
        self.x = pos.getX()
        self.y = pos.getY()
        self.w = w
        self.h = h
        self.label = label
        self.tooltip = "Input field for short text with newlines."
        self.wrap = True  # Equivalent behavior
        self.controller = None
        self.value = ""

    def setText(self, txt: str):
        self.value = txt

    def getText(self):
        return self.value

    def input_cb(self):
        # Placeholder for future callback logic, e.g., notify controller or update model
        pass

    def getRenderParams(self):
        return {
            "edit_x": self.x,
            "edit_y": self.y,
            "edit_w": self.w,
            "edit_h": self.h,
            "edit_label": self.label,
            "edit_value": self.value
        }
        
# ----------------------------- MODEL CLASS ------------------------------------

class Model:
    def __init__(self):
        self.lastChoice = 0
        self.lastInput = ""
        self.chView = None
        self.inpView = None
        self.factView = None
        self.sendPdfView = None

    def setLastChoice(self, ch):
        self.lastChoice = ch
        self.notify()

    def getLastChoice(self):
        return self.lastChoice

    def setChView(self, db: MyDisplayBox):
        self.chView = db

    def setInpView(self, db):
        self.inpView = db
    
    def setLastInput(self, txt):
        self.lastInput = txt

    def setFactView(self, db: MyDisplayBox):
        self.factView = db
    
    def setSendPdf(self, db: MyDisplayBox):
        self.sendPdfView = db

    def notify(self):
        if self.chView:
            self.chView.setText("Last choice is " + str(self.lastChoice))
        if self.inpView:
            self.inpView.setText(f"Last input is `{self.lastInput}`")

# ----------------------------- CONTROLLER CLASS -------------------------------

class Controller:
    def __init__(self):
        self.model = None

    def setModel(self, aModel: Model):
        self.model = aModel

    def chControl(self, aString: str): # apply the action from the GUI to the model
        try:
            ch = int(aString.strip().split()[-1])
            self.model.setLastChoice(ch)
        except Exception as e:
            print("Invalid input to Controller.chControl:", aString, e)
            
    def inpControl(self, first_input_box_text: str, second_input_box_text: str): # apply the action from the GUI to the model
        self.model.setLastInput(first_input_box_text)
        
        # Check if the 'export Excel' option (choice 1) is selected.
        if self.model.getLastChoice() == 1:
            self.handle_excel_export(first_input_box_text, second_input_box_text)
        
        #Compute based on choice and update view accordingly
        elif self.model.getLastChoice() == 2: # check if the option corresponds to export PDF
            try:
                self.handle_pdf_export(first_input_box_text, second_input_box_text)
            except Exception as e:
                self.model.inpView.setText("Invalid input")
                
        # Check if the 'send Excel' option (choice 3) is selected
        elif self.model.getLastChoice() == 3:
            self.handle_send_excel(first_input_box_text, second_input_box_text)
        
        # Check if the 'send PDF' option (choice 4) is selected
        elif self.model.getLastChoice() == 4:
            self.handle_send_pdf(first_input_box_text, second_input_box_text)
    
    
    def add_pdf_to_relevant_manager_archives(self, employee_email, pdf_path):
        """
        Adaugă PDF-ul la arhivele tuturor managerilor relevanți
        """
        try:
            global _manager_archives_in_progress
            
            emp = Employee.query.filter_by(email=employee_email).first()
            if not emp:
                return
            
            for manager_email, archive_info in _manager_archives_in_progress.items():
                manager = Employee.query.filter_by(email=manager_email).first()
                if not manager:
                    continue
                    
                # Verifică dacă angajatul aparține acestui manager sau este managerul însuși
                if emp.employee_id == manager.employee_id or emp.manager_id == manager.employee_id:
                    if add_pdf_to_manager_archive(manager_email, pdf_path):
                        archive_info['received_pdfs'].add(employee_email)
                        print(f"Added PDF for {employee_email} to manager {manager_email} archive")
                    
        except Exception as e:
            print(f"Error adding PDF to manager archives: {e}")

    def check_and_finalize_complete_archives(self):
        """
        Verifică și finalizează doar arhivele care sunt complete
        """
        try:
            global _manager_archives_in_progress, _manager_expected_employees, _pending_excel_for_managers
            
            managers_to_finalize = []
            
            for manager_email, archive_info in _manager_archives_in_progress.items():
                if manager_email in _manager_expected_employees:
                    expected_employees = _manager_expected_employees[manager_email]
                    received_pdfs = archive_info['received_pdfs']
                    
                    print(f"Manager {manager_email}:")
                    print(f"  Expected: {expected_employees}")
                    print(f"  Received: {received_pdfs}")
                    
                    # Verifică dacă toți angajații așteptați au primit PDF-uri
                    if expected_employees.issubset(received_pdfs):
                        print(f"  Archive is complete! Marking for finalization.")
                        managers_to_finalize.append(manager_email)
                    else:
                        missing = expected_employees - received_pdfs
                        print(f"  Still waiting for: {missing}")
            
            # Finalizează arhivele complete
            for manager_email in managers_to_finalize:
                if finalize_manager_archive(manager_email):
                    print(f"Archive successfully finalized for manager {manager_email}")
                    # Curăță tracking-ul
                    if manager_email in _manager_expected_employees:
                        del _manager_expected_employees[manager_email]
                    if manager_email in _pending_excel_for_managers:
                        del _pending_excel_for_managers[manager_email]
                else:
                    print(f"Failed to finalize archive for manager {manager_email}")
                    
        except Exception as e:
            print(f"Error checking and finalizing archives: {e}")

    def handle_send_pdf(self, input_text, second_input_text):
        """
        Versiunea îmbunătățită care verifică complet înainte de finalizare
        """
        try:
            lines = input_text.strip().split('\n')
            pdf_path = lines[0].strip() if lines else ""
            if not pdf_path or not os.path.exists(pdf_path) or not pdf_path.lower().endswith('.pdf'):
                self.model.sendPdfView.setText("Error: File invalid or not found")
                return

            emails = [e.strip() for e in second_input_text.splitlines() if e.strip()]
            if not emails:
                self.model.sendPdfView.setText("Error: No email addresses provided")
                return

            results, sent_files = [], []
            successful_sends = 0
            successful_recipients = []

            # Trimite PDF-urile către toți destinatarii
            for email in emails:
                emp = Employee.query.filter_by(email=email).first()
                if not emp or not emp.cnp:
                    results.append(f"{email}: Employee or CNP not found")
                    continue

                # Creează PDF criptat
                reader = PdfReader(pdf_path)
                writer = PdfWriter()
                for p in reader.pages:
                    writer.add_page(p)
                writer.encrypt(user_password=emp.cnp, owner_password=emp.cnp)

                base = os.path.splitext(os.path.basename(pdf_path))[0]
                enc_name = f"{base}_encrypted_{emp.employee_id}.pdf"
                encrypted_path = os.path.join(os.path.dirname(pdf_path), enc_name)
                
                with open(encrypted_path, 'wb') as out_f:
                    writer.write(out_f)

                # Trimite emailul
                sent = send_file_via_email(encrypted_path, [email])
                status = 'sent' if sent else 'failed'
                results.append(f"{email}: {status}")

                if sent:
                    successful_sends += 1
                    successful_recipients.append(email)
                    sent_files.append(encrypted_path)
                    
                    # Salvează în baza de date
                    db.session.add(ArchivedFile(
                        employee_id=emp.employee_id,
                        file_name=enc_name,
                        file_type='PDF',
                        path=encrypted_path,
                        sent_date=date.today()
                    ))

                    # Adaugă la arhivele managerilor relevanți
                    self.add_pdf_to_relevant_manager_archives(email, encrypted_path)

            # Commit la baza de date
            db.session.commit()

            # Verifică și finalizează arhivele complete
            self.check_and_finalize_complete_archives()

            # Afișează mesajul cu rezultatele
            if successful_recipients:
                receivers_text = ", ".join(successful_recipients[:3])
                if len(successful_recipients) > 3:
                    receivers_text += f" (+{len(successful_recipients)-3} more)"
                self.model.sendPdfView.setText(f"PDF sent to: {receivers_text}")
            else:
                self.model.sendPdfView.setText("No PDFs sent successfully")

        except Exception as e:
            self.model.sendPdfView.setText(f"Error: {e}")
            print(f"Error in handle_send_pdf_fixed: {e}")
    
    def handle_send_excel(self, input_text, second_input_text):
        """
        Versiunea actualizată pentru trimiterea Excel-ului
        """
        try:
            lines = input_text.strip().split('\n')
            file_path = lines[0].strip() if lines else ""
            if not file_path or not os.path.exists(file_path):
                self.model.factView.setText("Error: File not found")
                return

            recipient_emails = [e.strip() for e in second_input_text.splitlines() if e.strip()]
            if not recipient_emails:
                self.model.factView.setText("Error: No email addresses provided")
                return

            success = send_file_via_email(file_path, recipient_emails)
            if not success:
                self.model.factView.setText("Error: Failed to send email")
                return

            # Pentru fiecare manager din listă, folosește noua logică
            for email in recipient_emails:
                emp = Employee.query.filter_by(email=email).first()
                if emp and emp.role == 'manager':
                    check_and_create_archive_for_manager(email, file_path)

            # Salvează în arhivă
            for email in recipient_emails:
                emp = Employee.query.filter_by(email=email).first()
                db.session.add(ArchivedFile(
                    employee_id=emp.employee_id if emp else None,
                    file_name=os.path.basename(file_path),
                    file_type='Excel',
                    path=file_path,
                    sent_date=date.today()
                ))
            db.session.commit()

            # Afișează mesajul cu lista receiverilor
            receivers_text = ", ".join(recipient_emails[:3])
            if len(recipient_emails) > 3:
                receivers_text += f" (+{len(recipient_emails)-3} more)"
            
            self.model.factView.setText(f"Excel sent to: {receivers_text}")
            
        except Exception as e:
            self.model.factView.setText(f"Error: {e}")
    
    def handle_excel_export(self, input_text, second_input_text):
        """
        Gestioneaza exportul Excel bazat pe input-urile din GUI
        """
        try:
            lines = input_text.strip().split('\n')
            if not lines or not lines[0].strip():
                if self.model.chView:
                    self.model.chView.setText("Error: No folder path provided")
                return
            
            # The first line is the path of the folder.
            output_folder = lines[0].strip()
            
            # Colectam email-uri, daca exista
            if second_input_text.strip():
                email_addresses = [
                    e.strip() for e in second_input_text.splitlines() if e.strip()
                ]
            else:
                # daca e gola, vom lasa lista goala si helper-ul va prelua toate email-urile
                email_addresses = []
            
            # Make sure that the folder exists.
            os.makedirs(output_folder, exist_ok=True)
            
            # Generate the filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"employee_summary_{timestamp}.xlsx"
            output_path = os.path.join(output_folder, filename)
            
            # Collect the employees' data
            employees_data = get_employee_data_for_excel(email_addresses)
            
            if not employees_data:
                if self.model.chView:
                    self.model.chView.setText("Error: No employee data found")
                return
            
            # Create the Excel file
            success = create_excel_file(employees_data, output_path)
            
            if success:
                if self.model.chView:
                    self.model.chView.setText(f"{filename} created")
            else:
                if self.model.chView:
                    self.model.chView.setText("Error: Failed to create Excel file")
                    
        except Exception as e:
            if self.model.chView:
                self.model.chView.setText(f"Error: {str(e)}")
            print(f"Excel export error: {e}")
    
    def handle_pdf_export(self, input_text, second_input_text):
        """
        Generates the individual PDF reports with:
            personal details (name, surname, employee ID, CNP)
            salary for the current month
            
        Writes them to the folder given on the first line of input_text,
        reading e-mails either from second_input_text or from the
        remaining lines of input_text, and updates the “export PDF”
        display box with success or error.
        """
        try:
            # 1) determine output folder
            lines = input_text.strip().split("\n")
            if not lines or not lines[0].strip():
                self.model.inpView.setText("Error: No folder path provided")
                return
            output_folder = lines[0].strip()
            os.makedirs(output_folder, exist_ok=True)

            # 2) collect e‑mails
            block = second_input_text.strip() or "\n".join(lines[1:])
            emails = [e.strip() for e in block.split("\n") if e.strip()]
            if not emails:
                self.model.inpView.setText("Error: No email addresses provided")
                return

            # 3) fetch the data and build PDF
            data = get_employee_data_for_pdf(emails)
            if not data:
                self.model.inpView.setText("Error: No employee data found")
                return

            timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename    = f"employee_summary_{timestamp}.pdf"
            output_path = os.path.join(output_folder, filename)

            if create_pdf_file(data, output_path):
                self.model.inpView.setText(f"{filename} created")
            else:
                self.model.inpView.setText("Error: Failed to create PDF file")

        except Exception as e:
            self.model.inpView.setText(f"Error: {e}")
            print(f"PDF export error: {e}")

# ----------------------------- VIEW-CONTROLLER ASSOCIATION --------------------

class MyRadioButton:
    _id_counter = 0

    def __init__(self, pos: Point, w: int, h: int, slabel: str):
        self.x = pos.getX()
        self.y = pos.getY()
        self.w = w
        self.h = h
        self.label = slabel
        self.tooltip = "Radio button, only one button is set at a time."
        self.down_box = "FL_ROUND_DOWN_BOX"
        self.id = f"radio{MyRadioButton._id_counter}"
        MyRadioButton._id_counter += 1
        self.controller = None

    def getRenderParams(self):
        return {
            "id": self.id,
            "label": self.label
        }

    def setController(self, aCntrl):
        self.controller = aCntrl

    def radio_button_cb(self):
        if self.controller:
            self.controller.chControl(self.label)

class MyRadioGroup:
    def __init__(self, pos: Point, w: int, h: int, label: str, no: int):
        self.elts = []
        bpos = Point(pos.getX(), pos.getY())
        for i in range(no):
            bpos.setY(pos.getY() + i * 30)
            rb = MyRadioButton(bpos, w, h // no, f"My Choice {i + 1}")
            self.elts.append(rb)

    def getButtons(self):
        return self.elts
    
    def setController(self, aCntrl):
        for rb in self.elts:
            rb.setController(aCntrl)

# ----------------------------- CONNECTION LOGIC ----------------------------------------

class MyWindow:
    def __init__(self, pos: Point, w: int, h: int, title: str):
        if pos is None:
            self.x, self.y = 100, 200
        else:
            self.x, self.y = pos.getX(), pos.getY()
        self.w = w
        self.h = h
        self.title = title
        self.display_box = None
        self.return_button = None
        self.radio_buttons = []
        
        # Store references to all three display boxes
        self.firstdb = None
        self.seconddb = None
        self.thirddb = None
        self.fourthdb = None

    def addDisplayBox(self, display_box: MyDisplayBox):
        if self.firstdb is None:
            self.firstdb = display_box
        elif self.seconddb is None:
            self.seconddb = display_box
        elif self.thirddb is None:
            self.thirddb = display_box
        elif self.fourthdb is None:
            self.fourthdb = display_box

    def addReturnButton(self, return_button: MyReturnButton):
        self.return_button = return_button
        
    def addRadioButton(self, rb: MyRadioButton):
        self.radio_buttons.append(rb)
    
    def addRadioGroup(self, group):
        self.radio_buttons.extend(group.getButtons())

    def getRenderParams(self):
        params = {
            "x": self.x,
            "y": self.y,
            "w": self.w,
            "h": self.h,
            "title": self.title,
            "first_display_box_text": self.firstdb.getText() if self.firstdb else "",
            "second_display_box_text": self.seconddb.getText() if self.seconddb else "",
            "third_display_box_text": self.thirddb.getText() if self.thirddb else "",
            "fourth_display_box_text": self.thirddb.getText() if self.thirddb else "",
            "label": self.display_box.label if self.display_box else ""
        }
        if self.return_button:
            params.update(self.return_button.getRenderParams())
        if self.radio_buttons:
            params["radio_buttons"] = [rb.getRenderParams() for rb in self.radio_buttons]
        return params

# ----------------------------- ROUTING --------------------------

@app.route("/", methods=["GET", "POST"])
def index():
    
    # if user/admin not loggin, then send to /login
    if "user" not in session:
        return redirect(url_for("login"))
    
    
    posMainWindow = Point(100, 200)
    mainwindow = MyWindow(posMainWindow, 1150, 400, "Main Window")

    # Display Boxes
    posFirstDB = Point(100, 50)
    firstdb = MyDisplayBox(posFirstDB, 200, 50, "My display box")

    posSndDB = Point(360, 50)
    seconddb = MyDisplayBox(posSndDB, 200, 50, "Second display")

    posTrdDB = Point(620, 50)
    thirddb = MyDisplayBox(posTrdDB, 200, 50, "Third display")
    
    posFrthDB = Point(880, 50)
    fourthdb = MyDisplayBox(posFrthDB, 200, 50, "Fourth display")

    firstdb.setText("My first output text.")
    seconddb.setText("My second output text.")
    thirddb.setText("My third output text.")
    fourthdb.setText("My fourth output text.")

    mainwindow.addDisplayBox(firstdb)
    mainwindow.addDisplayBox(seconddb)
    mainwindow.addDisplayBox(thirddb)
    mainwindow.addDisplayBox(fourthdb)

    # Model and Controller
    model = Model()
    model.setChView(firstdb)    # Set firstdb for displaying Excel export messages
    model.setInpView(seconddb)  # Set seconddb for PDF export messages (choice 2)
    model.setFactView(thirddb)  # set the factorial view to the third display box
    model.setSendPdf(fourthdb)  # Set fourthdb for send PDF messages (choice 4)

    chCntrl = Controller()
    chCntrl.setModel(model)

    # Radio Group
    posRG = Point(160, 150)
    rg = MyRadioGroup(posRG, 150, 90, "MyChoice", 4)
    rg.setController(chCntrl)
    mainwindow.addRadioGroup(rg)

    # Return Button
    posRet = Point(400, 350)
    ret = MyReturnButton(posRet, 100, 25)
    mainwindow.addReturnButton(ret)

    # Edit Box input
    posEB = Point(400, 130)
    eb = MyEditBox(posEB, 150, 100, "&My Input")
    eb.setText("Initial edit text\nSecond line")

    # Edit Second Box input 
    posEB2 = Point(620, 130) # 400 px + 220 px = 620 px
    eb2 = MyEditBox(posEB2, 150, 100, "&My Input 2")
    eb2.setText("")

    if request.method == "POST":
        input_text = request.form.get("edit_box", "")
        second_text_input = request.form.get("edit_box2", "")
        selected_choice = request.form.get("radio_option", "") # processing the value of the selected radio button
        
        # Truncam fiecare linie la maxim 256 de caractere
        input_text = "\n".join([line[:256] for line in input_text.splitlines()])
        second_text_input = "\n".join([line[:256] for line in second_text_input.splitlines()])
        
        try:
            selected_choice_int = int(selected_choice)
        except Exception:
            selected_choice_int = 0
        
        # Set the choice in model and process both inputs through inpControl
        model.setLastChoice(selected_choice_int)
        
        # Process input text for both boxes
        chCntrl.inpControl(input_text, second_text_input)
        
        # Set current_input for rendering
        current_input = input_text
        eb.setText(input_text)
        
        # Set current_second_input for rendering
        current_second_input = second_text_input
        eb2.setText(second_text_input)
    else:
        # For GET requests, use the initial text
        current_input = eb.getText()
        current_second_input = eb2.getText()
        selected_choice = ""
    
    # Prepare render parameters
    render_params = mainwindow.getRenderParams()
    render_params.update({
        "selected_choice": int(selected_choice) if selected_choice else 0,
        "current_input": current_input,
        "current_second_input": current_second_input,
        "first_text": firstdb.getText(),
        "second_text": seconddb.getText(),
        "third_text": thirddb.getText(),
        "fourth_text": fourthdb.getText()
    })
    
    return render_template_string(HTML_TEMPLATE, **render_params)
    
if __name__ == "__main__":
    app.run(debug=True)