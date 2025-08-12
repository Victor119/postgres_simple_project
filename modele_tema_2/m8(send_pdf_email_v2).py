import base64
from base64 import b64encode
import calendar
from datetime import datetime, date
from dotenv import load_dotenv
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formataddr
from email import encoders
from flask import Flask, jsonify, render_template, render_template_string, request, redirect, session, url_for 
from flask_sqlalchemy import SQLAlchemy
import json
import logging
import mailersend 
from mailersend import emails
import msal
from O365 import Account, FileSystemTokenBackend
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from pypdf import PdfReader, PdfWriter
import pandas as pd
import requests
import smtplib
import ssl
import subprocess
import time
from typing import List
import uuid

app = Flask(__name__)

# Foloseste un secret key aleator la fiecare pornire, astfel incat sesiunile vechi sa fie invalidate
app.secret_key = os.environ.get('SECRET_KEY') or os.urandom(24)

#DB configuration
app.config['SQLALCHEMY_DATABASE_URI'] = (os.environ.get('DATABASE_URL') or 'postgresql://postgres:postgres@localhost:5432/master')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ----------------------------- HTML TEMPLATE ----------------------------------

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{{ title }}</title>
    <style>
        body {
            margin: 0;
            padding: 0;
            background-color: #f0f0f0;
        }
        .window {
            position: absolute;
            left: {{ x }}px;
            top: {{ y }}px;
            width: {{ w }}px;
            height: {{ h }}px;
            border: 2px solid #333;
            background-color: #ccc;
            box-shadow: 2px 2px 10px rgba(0,0,0,0.5);
            padding: 10px;
        }
        .display-container {
            position: absolute;
            top: 50px;
            left: 100px;
            display: flex;
            align-items: center;
        }
        .display-label {
            font-family: sans-serif;
            font-size: 14px;
            margin-right: 10px;
        }
        .display-box {
            width: 200px;
            height: 50px;
            border: 1px solid #666;
            background-color: #fff;
            font-family: monospace;
            font-size: 14px;
            padding: 5px;
            box-sizing: border-box;
        }
        .radio-button {
            position: absolute;
            top: 120px;
            left: 100px;
            font-family: sans-serif;
            font-size: 14px;
        }
        .return-button {
            position: absolute;
            left: {{ ret_x }}px;
            top: {{ ret_y }}px;
            width: {{ ret_w }}px;
            height: {{ ret_h }}px;
            font-size: 12px;
        }
        .radio-group {
            position: absolute;
            top: 150px;
            left: 160px;
            width: 200px;
            border: 1px solid #333;
            background-color: #eee;
            padding: 10px;
        }
    </style>
    <script>
        // Enforce max 256 characters per line and allow Enter key
        document.addEventListener('DOMContentLoaded', function() {
            function enforceMaxLineLength(textarea) {
                const lines = textarea.value.split('\n');
                for (let i = 0; i < lines.length; i++) {
                    if (lines[i].length > 256) {
                        lines[i] = lines[i].substring(0, 256);
                    }
                }
                textarea.value = lines.join('\n');
            }
            ['edit_box', 'edit_box2'].forEach(function(id) {
                const ta = document.getElementById(id);
                if (ta) {
                    ta.addEventListener('input', function() { enforceMaxLineLength(ta); });
                }
            });
        });
    </script>
</head>
<body>
    <div class="window">
        <div class="display-container">
            <div class="display-label">export Excel</div>
            <input class="display-box" type="text" value="{{ first_text }}" readonly>

            <div class="display-label" style="margin-left: 20px;">export PDF</div>
            <input class="display-box" type="text" value="{{ second_text }}" readonly>

            <div class="display-label" style="margin-left: 20px;">send Excel</div>
            <input class="display-box" type="text" value="{{ third_text }}" readonly>
            
            <div class="display-label" style="margin-left: 20px;">send PDF</div>
            <input class="display-box" type="text" value="{{ fourth_text }}" readonly>
        </div>
        
        <form method="post">
            <div class="radio-group">
                {% set custom_labels = ["export Excel", "export PDF", "send Excel", "send PDF"] %}
                {% for i in range(radio_buttons | length) %}
                <div class="radio-option">
                    <input type="radio" id="{{ radio_buttons[i].id }}" name="radio_option" value="{{ i + 1 }}" 
                        {% if selected_choice == i + 1 %}checked{% endif %}>
                    <label for="{{ radio_buttons[i].id }}">{{ custom_labels[i] }}</label>
                </div>
                {% endfor %}
            </div>
            
            <div style="position: absolute; top: 130px; left: 400px;">
                <label for="edit_box" style="font-family: sans-serif; font-size: 14px;">My Input</label><br>
                <textarea id="edit_box" name="edit_box" style="width: 150px; height: 100px; font-family: monospace; font-size: 14px;">{{ current_input }}</textarea>
            </div>
            
            <div style="position: absolute; top: 130px; left: 620px;">
                <label for="edit_box2" style="font-family: sans-serif; font-size: 14px;">My Input&nbsp;2</label><br>
                <textarea id="edit_box2" name="edit_box2" style="width: 150px; height: 100px; font-family: monospace; font-size: 14px;">{{ second_input }}</textarea>
            </div>
            
            <button class="return-button" type="submit">&Return</button>
        </form>
    </div>
</body>
</html>
"""

# ----------------------------- DATABASE Models ---------------------

# Employee model conform structurii existente
class Employee(db.Model):
    __tablename__ = 'employees'

    employee_id = db.Column(db.Integer,
                            primary_key=True)
    manager_id  = db.Column(db.Integer,
                            db.ForeignKey('employees.employee_id'),
                            nullable=True)
    role        = db.Column(db.String,      # 'manager' sau 'employee'
                            nullable=False)
    first_name  = db.Column(db.String(100))
    last_name   = db.Column(db.String(100))
    cnp         = db.Column(db.String(20),
                            unique=True,
                            nullable=False)
    username    = db.Column(db.String(64),
                            unique=True,
                            nullable=False)
    password    = db.Column(db.String(128),
                            nullable=False)
    email       = db.Column(db.String(120),
                            unique=True,
                            nullable=False)
    address     = db.Column(db.String(200))
    city        = db.Column(db.String(100))
    country     = db.Column(db.String(100))
    created_at  = db.Column(db.DateTime)

    @property
    def full_name(self):
        """Returnează numele complet al angajatului"""
        if self.first_name and self.last_name:
            return f"{self.first_name} {self.last_name}"
        elif self.first_name:
            return self.first_name
        elif self.last_name:
            return self.last_name
        else:
            return self.username or "N/A"

class Salary(db.Model):
    __tablename__ = 'salaries'
    salary_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    base_salary = db.Column(db.Numeric(10, 2))
    month = db.Column(db.Date)

class Bonus(db.Model):
    __tablename__ = 'bonuses'
    bonus_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    amount = db.Column(db.Numeric(10, 2))
    description = db.Column(db.String(255))
    month = db.Column(db.Date)

class Vacation(db.Model):
    __tablename__ = 'vacations'
    vacation_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    number_of_days = db.Column(db.Integer)
    reason = db.Column(db.String(255))

class WorkDay(db.Model):
    __tablename__ = 'work_days'
    work_day_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    month = db.Column(db.Date)
    number_of_days = db.Column(db.Integer)

class ArchivedFile(db.Model):
    __tablename__ = 'archived_files'
    file_id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.employee_id'))
    file_name = db.Column(db.String(255))
    file_type = db.Column(db.String(50))
    path = db.Column(db.String(500))
    sent_date = db.Column(db.Date)

# ----------------------------- HELPER FUNCTIONS FOR EXCEL ---------------------

def get_current_month_start():
    """Returneaza primul zi din luna curenta"""
    today = date.today()
    return date(today.year, today.month, 1)

def get_current_month_end():
    """ReturneazA ultima zi din luna curentA"""
    today = date.today()
    last_day = calendar.monthrange(today.year, today.month)[1]
    return date(today.year, today.month, last_day)

def get_employee_data_for_excel(email_addresses):
    """
    ColecteazA datele pentru toTi angajaTii cu adresele de email specificate
    pentru luna curenta
    """
    try:
        employees_data = []
        current_month_start = get_current_month_start()
        current_month_end = get_current_month_end()
        
        for email in email_addresses:
            email = email.strip()
            if not email:
                continue
                
            # Find the employee by email
            employee = Employee.query.filter_by(email=email).first()
            if not employee:
                continue
            
            # Get the salary for the current month
            salary = Salary.query.filter(
                Salary.employee_id == employee.employee_id,
                Salary.month >= current_month_start,
                Salary.month <= current_month_end
            ).first()
            
            # Get the working days for the current month
            work_days = WorkDay.query.filter(
                WorkDay.employee_id == employee.employee_id,
                WorkDay.month >= current_month_start,
                WorkDay.month <= current_month_end
            ).first()
            
            # Get the vacation days for the current month
            vacation_days = db.session.query(db.func.sum(Vacation.number_of_days)).filter(
                Vacation.employee_id == employee.employee_id,
                Vacation.start_date >= current_month_start,
                Vacation.end_date <= current_month_end
            ).scalar() or 0
            
            # Get the bonuses for the current month
            bonuses = Bonus.query.filter(
                Bonus.employee_id == employee.employee_id,
                Bonus.month >= current_month_start,
                Bonus.month <= current_month_end
            ).all()
            
            # Prepare the data for Excel
            employee_info = {
                'name': employee.full_name,
                'email': employee.email,
                'salary': float(salary.base_salary) if salary and salary.base_salary else 0.0,
                'work_days': work_days.number_of_days if work_days else 0,
                'vacation_days': vacation_days,
                'bonuses': []
            }
            
            # Add the bonuses
            for bonus in bonuses:
                employee_info['bonuses'].append({
                    'amount': float(bonus.amount) if bonus.amount else 0.0,
                    'description': bonus.description or 'N/A'
                })
            
            employees_data.append(employee_info)
        
        return employees_data
    
    except Exception as e:
        print(f"Error collecting employee data: {e}")
        return []

def create_excel_file(employees_data, output_path):
    """
    Creează fișierul Excel cu datele angajaților
    """
    try:
        # Create the workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Summary"
        
        # Styles for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # The headers of the columns
        headers = [
            "Employee Name", 
            "Email", 
            "Base Salary", 
            "Working Days", 
            "Vacation Days", 
            "Total Bonuses", 
            "Bonus Details"
        ]
        
        # Add the headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add the employee data
        for row, employee in enumerate(employees_data, 2):
            ws.cell(row=row, column=1, value=employee['name'])
            ws.cell(row=row, column=2, value=employee['email'])
            ws.cell(row=row, column=3, value=employee['salary'])
            ws.cell(row=row, column=4, value=employee['work_days'])
            ws.cell(row=row, column=5, value=employee['vacation_days'])
            
            # Calculate the total bonuses
            total_bonuses = sum(bonus['amount'] for bonus in employee['bonuses'])
            ws.cell(row=row, column=6, value=total_bonuses)
            
            # Add the details of the bonuses
            bonus_details = "; ".join([f"{bonus['description']}: {bonus['amount']}" for bonus in employee['bonuses']])
            ws.cell(row=row, column=7, value=bonus_details if bonus_details else "No bonuses")
        
        # Adjust the width of the columns
        column_widths = [20, 25, 15, 15, 15, 15, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Save the file
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False

# ----------------------------- HELPER FUNCTIONS FOR PDF ---------------------

def get_employee_data_for_pdf(email_addresses):
    """
    Collects the data required for the PDF export.
    Only personal details (first & last name, employee ID, CNP) and
    the salary due for the current month are returned.
    """
    try:
        employees_data = []
        current_month_start = get_current_month_start()
        current_month_end   = get_current_month_end()

        for email in (e.strip() for e in email_addresses):
            if not email:
                continue

            employee = Employee.query.filter_by(email=email).first()
            if not employee:
                continue

            salary_rec = Salary.query.filter(
                Salary.employee_id == employee.employee_id,
                Salary.month        >= current_month_start,
                Salary.month        <= current_month_end
            ).first()

            employees_data.append({
                "employee_id":   employee.employee_id,
                "first_name":    employee.first_name or "",
                "last_name":     employee.last_name  or "",
                "cnp":           employee.cnp        or "",
                "salary":        float(salary_rec.base_salary) if salary_rec and salary_rec.base_salary else 0.0,
            })

        return employees_data
    except Exception as e:
        print(f"Error collecting PDF data: {e}")
        return []

def create_pdf_file(employees_data, output_path):
    """
    Creates one aggregated PDF containing one short section
    for every employee in `employees_data`.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen       import canvas

        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        y = height - 40

        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, y, "Monthly Salary Report")
        y -= 30

        for emp in employees_data:
            if y < 120:          # simple page‑break
                c.showPage()
                y = height - 40
                c.setFont("Helvetica-Bold", 16)
                c.drawString(40, y, "Monthly Salary Report (cont.)")
                y -= 30

            c.setFont("Helvetica-Bold", 12)
            full_name = f"{emp['first_name']} {emp['last_name']}".strip()
            c.drawString(40, y, f"Employee: {full_name}")
            y -= 15

            c.setFont("Helvetica", 11)
            c.drawString(60, y, f"Employee ID: {emp['employee_id']}")
            y -= 15
            c.drawString(60, y, f"CNP: {emp['cnp']}")
            y -= 15
            c.drawString(60, y, f"Salary for current month: {emp['salary']:.2f}")
            y -= 25

        c.save()
        return True
    except Exception as e:
        print(f"Error creating PDF file: {e}")
        return False

# ----------------------------- EMAIL HELPER FUNCTION ---------------------

def send_file_via_email(
    file_path: str,
    recipient_emails: List[str],
    client_id: str = "f69f93d2-7437-462b-816d-42cd273538d0",
    tenant_id: str = "0b3fc178-b730-4e8b-9843-e81259237b77"
) -> bool:
    """
    Trimite email folosind delegated permissions - nu necesită admin consent
    Utilizatorul se va autentifica interactiv în browser
    """
    
    #1. Configure the MSAL application for delegated permissions
    app = msal.PublicClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}"
    )
    
    # Goals for delegated permissions (do not require admin consent)
    scopes = ["https://graph.microsoft.com/Mail.Send"]
    
    # 2. Try to get the token from the cache.
    accounts = app.get_accounts()
    result = None
    
    if accounts:
        #Try to obtain a silent token for the first account.
        result = app.acquire_token_silent(scopes, account=accounts[0])
    
    if not result:
        #If there is no token in the cache, request interactive authentication.
        print("Se deschide browser-ul pentru autentificare...")
        result = app.acquire_token_interactive(scopes)
    
    if "access_token" not in result:
        print(f"Eroare la autentificare: {result.get('error_description')}")
        return False
    
    token = result["access_token"]
    print("Autentificare reusita!")
    
    # 3. Read the file
    try:
        with open(file_path, 'rb') as f:
            file_content = f.read()
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        filename = os.path.basename(file_path)
    except Exception as ex:
        print(f"Eroare la citirea fișierului: {ex}")
        return False
    
    # 4. Build the payload
    recipients = [{"emailAddress": {"address": email}} for email in recipient_emails]
    
    email_payload = {
        "message": {
            "subject": "Employee Summary Report",
            "body": {
                "contentType": "HTML",
                "content": """
                <p>Hello,</p>
                <p>Please find attached the <strong>Employee Summary Report</strong>.</p>
                <p>Best regards,<br/>HR Department</p>
                """
            },
            "toRecipients": recipients,
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": filename,
                    "contentType": "application/octet-stream",
                    "contentBytes": file_base64
                }
            ]
        }
    }
    
    # 5. Send the email
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    
    graph_url = "https://graph.microsoft.com/v1.0/me/sendMail"
    
    try:
        response = requests.post(graph_url, json=email_payload, headers=headers)
        
        if response.status_code == 202:
            print("Email trimis cu succes!")
            return True
        else:
            print(f"Eroare: {response.status_code}")
            print(f"Răspuns: {response.text}")
            return False
            
    except Exception as e:
        print(f"Eroare la trimiterea email-ului: {e}")
        return False


# ----------------------------- API ENDPOINTS --------------------------------

@app.route("/createAggregatedEmployeeData", methods=["POST"])
def create_aggregated_employee_data():
    """
    API endpoint pentru generarea raportului Excel cu datele angajaților
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        email_addresses = data.get('emails', [])
        output_folder = data.get('output_folder', 'app/generated_excel')
        
        if not email_addresses:
            return jsonify({'error': 'No email addresses provided'}), 400
        
        # make sure folder exists
        os.makedirs(output_folder, exist_ok=True)
        
        # generate file name with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"employee_summary_{timestamp}.xlsx"
        output_path = os.path.join(output_folder, filename)
        
        # collect data about employee
        employees_data = get_employee_data_for_excel(email_addresses)
        
        if not employees_data:
            return jsonify({'error': 'No employee data found for provided emails'}), 404
        
        # create excel file
        success = create_excel_file(employees_data, output_path)
        
        if success:
            return jsonify({
                'success': True,
                'filename': filename,
                'path': output_path,
                'message': f'{filename} created successfully'
            }), 200
        else:
            return jsonify({'error': 'Failed to create Excel file'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

@app.route("/createPdfForEmployees", methods=["POST"])
def create_pdf_for_employees():
    """
    REST endpoint that receives a JSON body:
        { "emails": ["a@b.com", …], "output_folder": "app/generated_pdf" }
    and returns a single aggregated PDF report.
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data provided"}), 400

        email_addresses = data.get("emails", [])
        output_folder   = data.get("output_folder", "app/generated_pdf")

        if not email_addresses:
            return jsonify({"error": "No email addresses provided"}), 400

        os.makedirs(output_folder, exist_ok=True)
        employees_data = get_employee_data_for_pdf(email_addresses)
        if not employees_data:
            return jsonify({"error": "No employee data found for provided emails"}), 404

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"employee_summary_{timestamp}.pdf"
        output_path = os.path.join(output_folder, filename)

        if create_pdf_file(employees_data, output_path):
            return jsonify({
                "success": True,
                "filename": filename,
                "path": output_path,
                "message": f"{filename} created successfully"
            }), 200
        else:
            return jsonify({"error": "Failed to create PDF file"}), 500

    except Exception as e:
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route("/sendAggregatedEmployeeData", methods=["POST"])
def send_aggregated_employee_data():
    """
    API endpoint pentru trimiterea prin email a raportului Excel
    Primeste JSON cu: 
    {
        "file_path": "app/generated_excel/employee_summary_timestamp.xlsx",
        "emails": ["manager@acme.com", "hr@acme.com"]
    }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        file_path = data.get('file_path', '')
        recipient_emails = data.get('emails', [])
        
        if not file_path:
            return jsonify({'error': 'No file path provided'}), 400
        
        if not recipient_emails:
            return jsonify({'error': 'No recipient emails provided'}), 400
        
        # Check if the file exists
        if not os.path.exists(file_path):
            return jsonify({'error': f'File not found: {file_path}'}), 404
        
        # Send the email
        success = send_file_via_email(file_path, recipient_emails)
        
        if success:
            # Save information in the database for archive
            try:
                filename = os.path.basename(file_path)
                for email in recipient_emails:
                    # Find employee_id by email (optional)
                    employee = Employee.query.filter_by(email=email).first()
                    employee_id = employee.employee_id if employee else None
                    
                    archived_file = ArchivedFile(
                        employee_id=employee_id,
                        file_name=filename,
                        file_type='Excel',
                        path=file_path,
                        sent_date=date.today()
                    )
                    db.session.add(archived_file)
                
                db.session.commit()
            except Exception as e:
                print(f"Error saving to archive: {e}")
                # We are not returning an error because the email was sent successfully
            
            return jsonify({
                'success': True,
                'message': f'Excel file sent successfully to {len(recipient_emails)} recipients',
                'file_path': file_path,
                'recipients': recipient_emails
            }), 200
        else:
            return jsonify({'error': 'Failed to send email'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

@app.route("/sendPdfToEmployees", methods=["POST"])
def send_pdf_to_employees():
    """
    API endpoint pentru trimiterea PDF-urilor protejate cu parola (CNP) catre angajati.
    Asteapta JSON:
    {
        "file_path": "app/generated_pdf/employee_summary_....pdf",
        "emails": ["a@b.com", "c@d.com"]
    }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400

        pdf_path = data.get('file_path', '').strip()
        emails   = data.get('emails', [])
        if not pdf_path:
            return jsonify({'error': 'No file path provided'}), 400
        if not pdf_path.lower().endswith('.pdf'):
            return jsonify({'error': 'Provided file is not a PDF'}), 400
        if not emails:
            return jsonify({'error': 'No recipient emails provided'}), 400
        if not os.path.exists(pdf_path):
            return jsonify({'error': f'File not found: {pdf_path}'}), 404

        results = []
        for email in emails:
            emp = Employee.query.filter_by(email=email).first()
            if not emp or not emp.cnp:
                results.append({'email': email, 'status': 'Employee or CNP not found'})
                continue

            # encryption
            reader = PdfReader(pdf_path)
            writer = PdfWriter()
            for p in reader.pages:
                writer.add_page(p)
            writer.encrypt(user_pwd=emp.cnp)

            encrypted_path = f"{os.path.splitext(pdf_path)[0]}_{emp.employee_id}.pdf"
            with open(encrypted_path, 'wb') as f_out:
                writer.write(f_out)

            # send
            sent = send_file_via_email(encrypted_path, [email])
            status = 'sent' if sent else 'failed'
            results.append({'email': email, 'status': status})

            if sent:
                arch = ArchivedFile(
                    employee_id=emp.employee_id,
                    file_name=os.path.basename(encrypted_path),
                    file_type='PDF',
                    path=encrypted_path,
                    sent_date=date.today()
                )
                db.session.add(arch)

        db.session.commit()
        return jsonify({'results': results}), 200

    except Exception as e:
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        # Cautam direct in tabela employees
        user = Employee.query.filter_by(
            username=username,
            password=password
        ).first()

        if user:
            session["user"] = user.username
            return redirect(url_for("index"))
        else:
            error = "Credentiale invalide"
            return render_template("login.html", error=error)

    # GET: afisam formularul
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("user", None)     # sterge cheia din sesiune
    return redirect(url_for("login"))

# ----------------------------- CLASSES ---------------------------------------

class Point:
    def __init__(self, x, y):
        self._x = x
        self._y = y

    def getX(self):
        return self._x

    def getY(self):
        return self._y
    
    def setY(self, y):
        self._y = y

class Fl_Output:
    def __init__(self, x, y, w, h, label=None):
        self.x = x
        self.y = y
        self.width = w
        self.height = h
        self.label = label
        self._value = ""

    def value(self, txt):
        self._value = txt

    def redraw(self):
        pass

    def getText(self):
        return self._value

class MyDisplayBox(Fl_Output):
    def __init__(self, pos: Point, w: int, h: int, label: str = None):
        super().__init__(pos.getX(), pos.getY(), w, h, label)

    def setText(self, txt: str):
        self.value(txt)
        self.redraw()

class MyReturnButton:
    def __init__(self, pos: Point, w: int, h: int, label: str = "&Return"):
        self.x = pos.getX()
        self.y = pos.getY()
        self.w = w
        self.h = h
        self.label = label
        self.tooltip = "Push Return button to exit"
        self.labelsize = 12

    def getRenderParams(self):
        return {
            "ret_x": self.x,
            "ret_y": self.y,
            "ret_w": self.w,
            "ret_h": self.h,
            "label": self.label
        }

# ----------------------------- EDIT BOX CLASS ----------------------------------

class MyEditBox:
    def __init__(self, pos: Point, w: int, h: int, label: str):
        self.x = pos.getX()
        self.y = pos.getY()
        self.w = w
        self.h = h
        self.label = label
        self.tooltip = "Input field for short text with newlines."
        self.wrap = True  # Equivalent behavior
        self.controller = None
        self.value = ""

    def setText(self, txt: str):
        self.value = txt

    def getText(self):
        return self.value

    def input_cb(self):
        # Placeholder for future callback logic, e.g., notify controller or update model
        pass

    def getRenderParams(self):
        return {
            "edit_x": self.x,
            "edit_y": self.y,
            "edit_w": self.w,
            "edit_h": self.h,
            "edit_label": self.label,
            "edit_value": self.value
        }
        
# ----------------------------- MODEL CLASS ------------------------------------

class Model:
    def __init__(self):
        self.lastChoice = 0
        self.lastInput = ""
        self.chView = None
        self.inpView = None
        self.factView = None
        self.sendPdfView = None

    def setLastChoice(self, ch):
        self.lastChoice = ch
        self.notify()

    def getLastChoice(self):
        return self.lastChoice

    def setChView(self, db: MyDisplayBox):
        self.chView = db

    def setInpView(self, db):
        self.inpView = db
    
    def setLastInput(self, txt):
        self.lastInput = txt

    def setFactView(self, db: MyDisplayBox):
        self.factView = db
    
    def setSendPdf(self, db: MyDisplayBox):
        self.sendPdfView = db

    def notify(self):
        if self.chView:
            self.chView.setText("Last choice is " + str(self.lastChoice))
        if self.inpView:
            self.inpView.setText(f"Last input is `{self.lastInput}`")

# ----------------------------- CONTROLLER CLASS -------------------------------

class Controller:
    def __init__(self):
        self.model = None

    def setModel(self, aModel: Model):
        self.model = aModel

    def chControl(self, aString: str): # apply the action from the GUI to the model
        try:
            ch = int(aString.strip().split()[-1])
            self.model.setLastChoice(ch)
        except Exception as e:
            print("Invalid input to Controller.chControl:", aString, e)
            
    def inpControl(self, first_input_box_text: str, second_input_box_text: str): # apply the action from the GUI to the model
        self.model.setLastInput(first_input_box_text)
        
        # Check if the 'export Excel' option (choice 1) is selected.
        if self.model.getLastChoice() == 1:
            self.handle_excel_export(first_input_box_text, second_input_box_text)
        
        #Compute based on choice and update view accordingly
        elif self.model.getLastChoice() == 2: # check if the option corresponds to export PDF
            try:
                self.handle_pdf_export(first_input_box_text, second_input_box_text)
            except Exception as e:
                self.model.inpView.setText("Invalid input")
                
        # Check if the 'send Excel' option (choice 3) is selected
        elif self.model.getLastChoice() == 3:
            self.handle_send_excel(first_input_box_text, second_input_box_text)
        
        # Check if the 'send PDF' option (choice 4) is selected
        elif self.model.getLastChoice() == 4:
            self.handle_send_pdf(first_input_box_text, second_input_box_text)
    
    def handle_send_pdf(self, input_text, second_input_text):
        """
        Gestioneaza trimiterea prin email a fisierelor PDF.
        Prima linie din input_text: path catre fisierul PDF.
        second_input_text: adresele de email (una per linie).
        Fiecare PDF va fi protejat cu parola = CNP-ul angajatului.
        """
        try:
            lines = input_text.strip().split('\n')
            if not lines or not lines[0].strip():
                if self.model.sendPdfView:
                    self.model.sendPdfView.setText("Error: No file path provided")
                return
            
            pdf_path = lines[0].strip()
            
            # Verifica daca fisierul exista si este PDF
            if not os.path.exists(pdf_path):
                if self.model.sendPdfView:
                    self.model.sendPdfView.setText("Error: File not found")
                return
            
            if not pdf_path.lower().endswith('.pdf'):
                if self.model.sendPdfView:
                    self.model.sendPdfView.setText("Error: File is not a PDF")
                return

            # Colecteaza email-urile din al doilea input box
            emails = [e.strip() for e in (second_input_text.strip().split('\n')) if e.strip()]
            if not emails:
                if self.model.sendPdfView:
                    self.model.sendPdfView.setText("Error: No email addresses provided")
                return

            results = []
            successful_sends = 0
            
            for email in emails:
                try:
                    # Gaseste angajatul si CNP-ul
                    emp = Employee.query.filter_by(email=email).first()
                    if not emp or not emp.cnp:
                        results.append(f"{email}: Employee or CNP not found")
                        continue

                    # Creeaza fisierul PDF criptat
                    reader = PdfReader(pdf_path)
                    writer = PdfWriter()
                    
                    # Adauga toate paginile
                    for page in reader.pages:
                        writer.add_page(page)
                    
                    # Cripteaza cu CNP-ul angajatului
                    writer.encrypt(user_password=emp.cnp, owner_password=emp.cnp)

                    # Creeaza path-ul pentru fisierul criptat
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    encrypted_filename = f"{base_name}_encrypted_{emp.employee_id}.pdf"
                    encrypted_path = os.path.join(os.path.dirname(pdf_path), encrypted_filename)
                    
                    # Salveaza fiaierul criptat
                    with open(encrypted_path, 'wb') as out_f:
                        writer.write(out_f)

                    # Verifica ca fisierul criptat a fost creat
                    if not os.path.exists(encrypted_path):
                        results.append(f"{email}: Failed to create encrypted PDF")
                        continue

                    # Trimite email-ul
                    print(f"Attempting to send {encrypted_path} to {email}")
                    success = send_file_via_email(encrypted_path, [email])
                    
                    if success:
                        successful_sends += 1
                        results.append(f"{email}: sent")
                        
                        # Arhiveaza daca a fost trimis cu succes
                        try:
                            arch = ArchivedFile(
                                employee_id=emp.employee_id,
                                file_name=encrypted_filename,
                                file_type='PDF',
                                path=encrypted_path,
                                sent_date=date.today()
                            )
                            db.session.add(arch)
                        except Exception as archive_error:
                            print(f"Archive error for {email}: {archive_error}")
                            
                    else:
                        results.append(f"{email}: failed to send")
                        
                except Exception as e:
                    print(f"Error processing {email}: {str(e)}")
                    results.append(f"{email}: Error - {str(e)}")

            # Commit la baza de date pentru arhivare
            try:
                db.session.commit()
            except Exception as commit_error:
                print(f"Database commit error: {commit_error}")
                db.session.rollback()

            # Actualizeaza view-ul cu rezultatele
            if self.model.sendPdfView:
                summary = f"Sent: {successful_sends}/{len(emails)}. Details: " + "; ".join(results[:3])  # Limiteaza output-ul
                if len(results) > 3:
                    summary += f"... and {len(results) - 3} more"
                self.model.sendPdfView.setText(summary)
                    
        except Exception as e:
            print(f"Send PDF error: {e}")
            if self.model.sendPdfView:
                self.model.sendPdfView.setText(f"Error: {str(e)}")


    # Funcție helper pentru debugging
    def debug_pdf_sending(pdf_path, email):
        """
        Funcție de debugging pentru a testa trimiterea PDF-urilor
        """
        print(f"=== DEBUG PDF SENDING ===")
        print(f"PDF Path: {pdf_path}")
        print(f"File exists: {os.path.exists(pdf_path)}")
        print(f"Email: {email}")
        
        # Verifică angajatul
        emp = Employee.query.filter_by(email=email).first()
        if emp:
            print(f"Employee found: {emp.full_name}, CNP: {emp.cnp}")
        else:
            print("Employee not found")
            return False
        
        try:
            # Test PDF reading
            reader = PdfReader(pdf_path)
            print(f"PDF pages: {len(reader.pages)}")
            
            # Test encryption
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            writer.encrypt(user_password=emp.cnp, owner_password=emp.cnp)
            
            test_path = f"test_encrypted_{emp.employee_id}.pdf"
            with open(test_path, 'wb') as f:
                writer.write(f)
            
            print(f"Encrypted PDF created: {test_path}")
            print(f"Encrypted file size: {os.path.getsize(test_path)} bytes")
            
            return True
            
        except Exception as e:
            print(f"Error in PDF processing: {e}")
            return False
    
    def handle_send_excel(self, input_text, second_input_text):
        """
        Gestioneaza trimiterea prin email a fisierului Excel
        Prima linie din input_text = path catre fisierul Excel
        second_input_text = email-urile destinatarilor (cate unul pe linie)
        """
        try:
            lines = input_text.strip().split('\n')
            if not lines or not lines[0].strip():
                if self.model.factView:  # Folosim factView pentru "send Excel"
                    self.model.factView.setText("Error: No file path provided")
                return
            
            # Prima linie este path-ul către fișierul Excel
            file_path = lines[0].strip()
            
            # Verifica daca fisierul exista
            if not os.path.exists(file_path):
                if self.model.factView:
                    self.model.factView.setText("Error: File not found")
                return
            
            # Colecteaza email-urile din al doilea input box
            if second_input_text.strip():
                email_lines = second_input_text.strip().split('\n')
                recipient_emails = [email.strip() for email in email_lines if email.strip()]
            else:
                if self.model.factView:
                    self.model.factView.setText("Error: No email addresses provided")
                return
            
            if not recipient_emails:
                if self.model.factView:
                    self.model.factView.setText("Error: No valid email addresses")
                return
            
            # Trimite fisierul prin email
            success = send_file_via_email(file_path, recipient_emails)
            
            if success:
                # Salveaza in arhiva
                try:
                    filename = os.path.basename(file_path)
                    for email in recipient_emails:
                        employee = Employee.query.filter_by(email=email).first()
                        employee_id = employee.employee_id if employee else None
                        
                        archived_file = ArchivedFile(
                            employee_id=employee_id,
                            file_name=filename,
                            file_type='Excel',
                            path=file_path,
                            sent_date=date.today()
                        )
                        db.session.add(archived_file)
                    
                    db.session.commit()
                except Exception as e:
                    print(f"Error saving to archive: {e}")
                
                if self.model.factView:
                    self.model.factView.setText(f"Excel sent to {len(recipient_emails)} recipients")
            else:
                if self.model.factView:
                    self.model.factView.setText("Error: Failed to send email")
                    
        except Exception as e:
            if self.model.factView:
                self.model.factView.setText(f"Error: {str(e)}")
            print(f"Send Excel error: {e}")
    
    def handle_excel_export(self, input_text, second_input_text):
        """
        Gestioneaza exportul Excel bazat pe input-urile din GUI
        """
        try:
            lines = input_text.strip().split('\n')
            if not lines or not lines[0].strip():
                if self.model.chView:
                    self.model.chView.setText("Error: No folder path provided")
                return
            
            # The first line is the path of the folder.
            output_folder = lines[0].strip()
            
            # If there is a second input box, use it for emails.
            if second_input_text.strip():
                email_lines = second_input_text.strip().split('\n')
                email_addresses = [email.strip() for email in email_lines if email.strip()]
            else:
                # Otherwise, the emails are on the following lines of the first input.
                email_addresses = [email.strip() for email in lines[1:] if email.strip()]
            
            if not email_addresses:
                if self.model.chView:
                    self.model.chView.setText("Error: No email addresses provided")
                return
            
            # Make sure that the folder exists.
            os.makedirs(output_folder, exist_ok=True)
            
            # Generate the filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"employee_summary_{timestamp}.xlsx"
            output_path = os.path.join(output_folder, filename)
            
            # Collect the employees' data
            employees_data = get_employee_data_for_excel(email_addresses)
            
            if not employees_data:
                if self.model.chView:
                    self.model.chView.setText("Error: No employee data found")
                return
            
            # Create the Excel file
            success = create_excel_file(employees_data, output_path)
            
            if success:
                if self.model.chView:
                    self.model.chView.setText(f"{filename} created")
            else:
                if self.model.chView:
                    self.model.chView.setText("Error: Failed to create Excel file")
                    
        except Exception as e:
            if self.model.chView:
                self.model.chView.setText(f"Error: {str(e)}")
            print(f"Excel export error: {e}")
    
    def handle_pdf_export(self, input_text, second_input_text):
        """
        Generates the individual PDF reports with:
            personal details (name, surname, employee ID, CNP)
            salary for the current month
            
        Writes them to the folder given on the first line of input_text,
        reading e-mails either from second_input_text or from the
        remaining lines of input_text, and updates the “export PDF”
        display box with success or error.
        """
        try:
            # 1) determine output folder
            lines = input_text.strip().split("\n")
            if not lines or not lines[0].strip():
                self.model.inpView.setText("Error: No folder path provided")
                return
            output_folder = lines[0].strip()
            os.makedirs(output_folder, exist_ok=True)

            # 2) collect e‑mails
            block = second_input_text.strip() or "\n".join(lines[1:])
            emails = [e.strip() for e in block.split("\n") if e.strip()]
            if not emails:
                self.model.inpView.setText("Error: No email addresses provided")
                return

            # 3) fetch the data and build PDF
            data = get_employee_data_for_pdf(emails)
            if not data:
                self.model.inpView.setText("Error: No employee data found")
                return

            timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename    = f"employee_summary_{timestamp}.pdf"
            output_path = os.path.join(output_folder, filename)

            if create_pdf_file(data, output_path):
                self.model.inpView.setText(f"{filename} created")
            else:
                self.model.inpView.setText("Error: Failed to create PDF file")

        except Exception as e:
            self.model.inpView.setText(f"Error: {e}")
            print(f"PDF export error: {e}")

# ----------------------------- VIEW-CONTROLLER ASSOCIATION --------------------

class MyRadioButton:
    _id_counter = 0

    def __init__(self, pos: Point, w: int, h: int, slabel: str):
        self.x = pos.getX()
        self.y = pos.getY()
        self.w = w
        self.h = h
        self.label = slabel
        self.tooltip = "Radio button, only one button is set at a time."
        self.down_box = "FL_ROUND_DOWN_BOX"
        self.id = f"radio{MyRadioButton._id_counter}"
        MyRadioButton._id_counter += 1
        self.controller = None

    def getRenderParams(self):
        return {
            "id": self.id,
            "label": self.label
        }

    def setController(self, aCntrl):
        self.controller = aCntrl

    def radio_button_cb(self):
        if self.controller:
            self.controller.chControl(self.label)

class MyRadioGroup:
    def __init__(self, pos: Point, w: int, h: int, label: str, no: int):
        self.elts = []
        bpos = Point(pos.getX(), pos.getY())
        for i in range(no):
            bpos.setY(pos.getY() + i * 30)
            rb = MyRadioButton(bpos, w, h // no, f"My Choice {i + 1}")
            self.elts.append(rb)

    def getButtons(self):
        return self.elts
    
    def setController(self, aCntrl):
        for rb in self.elts:
            rb.setController(aCntrl)

# ----------------------------- CONNECTION LOGIC ----------------------------------------

class MyWindow:
    def __init__(self, pos: Point, w: int, h: int, title: str):
        if pos is None:
            self.x, self.y = 100, 200
        else:
            self.x, self.y = pos.getX(), pos.getY()
        self.w = w
        self.h = h
        self.title = title
        self.display_box = None
        self.return_button = None
        self.radio_buttons = []
        
        # Store references to all three display boxes
        self.firstdb = None
        self.seconddb = None
        self.thirddb = None
        self.fourthdb = None

    def addDisplayBox(self, display_box: MyDisplayBox):
        if self.firstdb is None:
            self.firstdb = display_box
        elif self.seconddb is None:
            self.seconddb = display_box
        elif self.thirddb is None:
            self.thirddb = display_box
        elif self.fourthdb is None:
            self.fourthdb = display_box

    def addReturnButton(self, return_button: MyReturnButton):
        self.return_button = return_button
        
    def addRadioButton(self, rb: MyRadioButton):
        self.radio_buttons.append(rb)
    
    def addRadioGroup(self, group):
        self.radio_buttons.extend(group.getButtons())

    def getRenderParams(self):
        params = {
            "x": self.x,
            "y": self.y,
            "w": self.w,
            "h": self.h,
            "title": self.title,
            "first_display_box_text": self.firstdb.getText() if self.firstdb else "",
            "second_display_box_text": self.seconddb.getText() if self.seconddb else "",
            "third_display_box_text": self.thirddb.getText() if self.thirddb else "",
            "fourth_display_box_text": self.thirddb.getText() if self.thirddb else "",
            "label": self.display_box.label if self.display_box else ""
        }
        if self.return_button:
            params.update(self.return_button.getRenderParams())
        if self.radio_buttons:
            params["radio_buttons"] = [rb.getRenderParams() for rb in self.radio_buttons]
        return params

# ----------------------------- ROUTING --------------------------

@app.route("/", methods=["GET", "POST"])
def index():
    
    # if user/admin not loggin, then send to /login
    if "user" not in session:
        return redirect(url_for("login"))
    
    
    posMainWindow = Point(100, 200)
    mainwindow = MyWindow(posMainWindow, 1150, 400, "Main Window")

    # Display Boxes
    posFirstDB = Point(100, 50)
    firstdb = MyDisplayBox(posFirstDB, 200, 50, "My display box")

    posSndDB = Point(360, 50)
    seconddb = MyDisplayBox(posSndDB, 200, 50, "Second display")

    posTrdDB = Point(620, 50)
    thirddb = MyDisplayBox(posTrdDB, 200, 50, "Third display")
    
    posFrthDB = Point(880, 50)
    fourthdb = MyDisplayBox(posFrthDB, 200, 50, "Fourth display")

    firstdb.setText("My first output text.")
    seconddb.setText("My second output text.")
    thirddb.setText("My third output text.")
    fourthdb.setText("My fourth output text.")

    mainwindow.addDisplayBox(firstdb)
    mainwindow.addDisplayBox(seconddb)
    mainwindow.addDisplayBox(thirddb)
    mainwindow.addDisplayBox(fourthdb)

    # Model and Controller
    model = Model()
    model.setChView(firstdb)    # Set firstdb for displaying Excel export messages
    model.setInpView(seconddb)  # Set seconddb for PDF export messages (choice 2)
    model.setFactView(thirddb)  # set the factorial view to the third display box
    model.setSendPdf(fourthdb)  # Set fourthdb for send PDF messages (choice 4)

    chCntrl = Controller()
    chCntrl.setModel(model)

    # Radio Group
    posRG = Point(160, 150)
    rg = MyRadioGroup(posRG, 150, 90, "MyChoice", 4)
    rg.setController(chCntrl)
    mainwindow.addRadioGroup(rg)

    # Return Button
    posRet = Point(400, 350)
    ret = MyReturnButton(posRet, 100, 25)
    mainwindow.addReturnButton(ret)

    # Edit Box input
    posEB = Point(400, 130)
    eb = MyEditBox(posEB, 150, 100, "&My Input")
    eb.setText("Initial edit text\nSecond line")

    # Edit Second Box input 
    posEB2 = Point(620, 130) # 400 px + 220 px = 620 px
    eb2 = MyEditBox(posEB2, 150, 100, "&My Input 2")
    eb2.setText("")

    if request.method == "POST":
        input_text = request.form.get("edit_box", "")
        second_text_input = request.form.get("edit_box2", "")
        selected_choice = request.form.get("radio_option", "") # processing the value of the selected radio button
        
        # Truncam fiecare linie la maxim 256 de caractere
        input_text = "\n".join([line[:256] for line in input_text.splitlines()])
        second_text_input = "\n".join([line[:256] for line in second_text_input.splitlines()])
        
        try:
            selected_choice_int = int(selected_choice)
        except Exception:
            selected_choice_int = 0
        
        # Set the choice in model and process both inputs through inpControl
        model.setLastChoice(selected_choice_int)
        
        # Process input text for both boxes
        chCntrl.inpControl(input_text, second_text_input)
        
        # Set current_input for rendering
        current_input = input_text
        eb.setText(input_text)
        
        # Set current_second_input for rendering
        current_second_input = second_text_input
        eb2.setText(second_text_input)
    else:
        # For GET requests, use the initial text
        current_input = eb.getText()
        current_second_input = eb2.getText()
        selected_choice = ""
    
    # Prepare render parameters
    render_params = mainwindow.getRenderParams()
    render_params.update({
        "selected_choice": int(selected_choice) if selected_choice else 0,
        "current_input": current_input,
        "current_second_input": current_second_input,
        "first_text": firstdb.getText(),
        "second_text": seconddb.getText(),
        "third_text": thirddb.getText(),
        "fourth_text": fourthdb.getText()
    })
    
    return render_template_string(HTML_TEMPLATE, **render_params)
    
if __name__ == "__main__":
    app.run(debug=True)